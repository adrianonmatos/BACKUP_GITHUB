import time
import sys
import os

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

from openpyxl import load_workbook
import pandas as pd
import psycopg2


###################################### CHECK FOR ARGUMENTS ##########################################################
if (len(sys.argv) == 1):
    print("Necessario especificar lista de anos. ex.: 2021 2022")
    sys.exit(1)

####################################### CLEAN-UP DATABASE ########################################################
def conecta_db():
    con = psycopg2.connect(host='srv1.deinfe.tce.am.gov.br',
            database = 'dados_abertos',
            user = 'postgres',
            password = 'deinfeusrpos')
    return con

def execute_sql(sql):
    con = conecta_db()
    cur = con.cursor()
    cur.execute(sql)
    con.commit()
    con.close()

sql = 'DROP TABLE IF EXISTS saude.tbFNS'
execute_sql(sql)

sql = '''CREATE TABLE saude.tbFNS
    ( uf character varying(2),
      municipio character varying(32),
      entidade character varying(128),
      cpf_cnpj character varying(16),
      bloco character varying(128),
      grupo character varying(128),
      acao character varying(256),
      acao_detalhada character varying(256),
      competencia_parcela character varying(16),
      No_OB character varying(16),
      Data_OB character varying(16),
      banco character varying(8),
      agencia_OB character varying(8),
      conta_OB character varying(16),
      valor_total character varying(16),
      desconto character varying(16),
      valor_liquido character varying(16),
      processo character varying(32),
      No_proposta character varying(32)
    )'''

execute_sql(sql)
print("CLEAN-UP BASE DE DADOS")

###################################### DOWNLOAD FNS EXCEL TABLES  ##############################
options = webdriver.ChromeOptions()
#options.add_argument("--headless")
#options.add_argument("--no-sandbox")
#options.add_argument("--disable-dev-shm-usage")

#chrome_prefs = {"download.default_directory": r"/home/adrianomatos"}
#options.experimental_options["prefs"] = chrome_prefs

options.headless = True
driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = options)


driver.get("https://consultafns.saude.gov.br/#/consolidada")
time.sleep(2)

for num in range(len(sys.argv)-1):
    print("Ano: ", sys.argv[num+1])

    select_ano = driver.find_element(By.ID, "ano")
    select_ano_obj = Select(select_ano)
    select_ano_obj.select_by_visible_text(sys.argv[num+1])

    select_estado = driver.find_element(By.ID, "estado")    
    select_estado_obj = Select(select_estado)
    select_estado_obj.select_by_visible_text("AMAZONAS")

    select_municipio = driver.find_element(By.ID, "municipio")
    select_municipio_obj = Select(select_municipio)
    select_municipio_obj.select_by_visible_text("Todos")

    select_tipoRepasse = driver.find_element(By.ID, "tipoRepasse")
    select_tipoRepasse_obj = Select(select_tipoRepasse)
    select_tipoRepasse_obj.select_by_visible_text("Todos")

    TEMPO_MAX = 200
    WebDriverWait(driver, TEMPO_MAX).until(EC.invisibility_of_element_located(
        (By.CSS_SELECTOR,
        ".overlayLoading")
    ))
    WebDriverWait(driver, TEMPO_MAX).until(EC.presence_of_element_located(
        (By.XPATH,
        "/html/body/div[2]/section/div/div[3]/div/div[2]/div/form/div[2]/div/div/button[1]")
    )).click()
    
    print('AGUARDANDO CARREGAR DADOS!!!')
    
    WebDriverWait(driver, TEMPO_MAX).until(EC.invisibility_of_element_located(
        (By.CSS_SELECTOR,
        ".overlayLoading")
    ))
    WebDriverWait(driver, TEMPO_MAX).until(EC.presence_of_element_located(
        (By.XPATH,
        "/html/body/div[2]/section/div/div[3]/div/div[3]/div[2]/div[4]/div/div/button[1]")
    )).click()

    print("AGUARDANDO DOWNLOAD PLANILHA")
    time.sleep(25)

###################################### EXTRACT DATA #################################################
    download_dir = './'
    old_fileName = 'consulta-consolidada-planilha'
    new_fileName = sys.argv[num+1]
    extension = '.xlsx'

    if(os.path.exists(download_dir + new_fileName + extension)):
        os.remove(download_dir + new_fileName + extension)
    os.rename(download_dir + old_fileName + extension, download_dir + new_fileName + extension)
    print("RENOMEANDO ARQUIVOS")

    wb = load_workbook(download_dir + new_fileName + extension)
    sheet = wb.worksheets[0]

    row_count = sheet.max_row
    col_count = sheet.max_column

    def find_header_row_idx(ws):
        for row in range(1, row_count + 1):
            coordinate = "{}{}".format("A", row)
            if sheet[coordinate].value == "UF":
                return row
        return None

    header_row_idx = find_header_row_idx(sheet)
    first_data_row = header_row_idx + 1

######################################## INSERT DATA INTO DATABASE ##############################################
    def insert_row(sql):
        con = conecta_db()
        cur = con.cursor()
        try:
            cur.execute(sql)
            con.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            print("Error: %s" % error)
            print(sql)
            con.rollback()
            cur.close()
            return 1
        cur.close()

    print("INSERINDO REGISTROS NO DB")
    for row in sheet.iter_rows(min_row=first_data_row , max_col=col_count, max_row=row_count, values_only = True):
        if(row[0]!=None and row[1]!=None and row[2]!=None):
            sql = """
            INSERT into saude.tbFNS (uf, municipio, entidade, cpf_cnpj, bloco, grupo, acao, acao_detalhada, competencia_parcela, No_OB, Data_OB, banco, agencia_OB, conta_OB, valor_total, desconto, valor_liquido, processo, No_proposta)
            values('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s');
            """ % (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17], row[18])
            insert_row(sql)

######################################### DISPOSE RESOURCES ####################################################################
driver.quit()
